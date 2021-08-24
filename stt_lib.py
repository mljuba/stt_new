"""
@author: ljmarjanovic
"""
# import sys
# import time

import xlsxwriter
import pyodbc
import openpyxl as opx

import calendar
import datetime
from datetime import date
import kontrole_lib as kon

# import tkinter as tk
# from tkinter import filedialog
# from tkinter import messagebox
# from tkinter import ttk
# from tkinter.ttk import Progressbar
# from tkinter import Menu


def populateSTTSheet(sheet_name, dateto):
    # create sheet headers
    pomdate_god = int(kon.right(dateto, 4))
    pomdate_mes = int(kon.right(kon.left(dateto, 5), 2))
    pomdate_dan = int(kon.left(dateto, 2))

    pomdate = datetime.datetime(pomdate_god, pomdate_mes, pomdate_dan)

    if pomdate.month == 1:
        pomM = datetime.datetime(pomdate.year - 1, 12, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('H4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year - 1, 11, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('G4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year - 1, 10, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('F4', pomMes, cell_wrap)

    elif pomdate.month == 2:
        pomM = datetime.datetime(pomdate.year, pomdate.month - 1, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('H4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year - 1, 12, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('G4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year - 1, 11, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('F4', pomMes, cell_wrap)

    elif pomdate.month == 3:
        pomM = datetime.datetime(pomdate.year, pomdate.month - 1, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('H4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year, pomdate.month - 2, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('G4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year - 1, 12, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('F4', pomMes, cell_wrap)
    else:
        pomM = datetime.datetime(pomdate.year, pomdate.month - 1, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('H4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year, pomdate.month - 2, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('G4', pomMes, cell_wrap)

        pomM = datetime.datetime(pomdate.year, pomdate.month - 3, 1)
        pomMes = pomM.strftime("%B %Y")
        sheet_name.write('F4', pomMes, cell_wrap)

    if sheet_name.get_name() == "CP":
        spom = str(datetime.datetime.now().isocalendar()[1])
        sheet_name.write('B2', 'Weeks: ' + spom)
    else:
        sheet_name.write_formula('B2', '=CP!B2')

    sheet_name.write('B3', '01-' + dateto)
    sheet_name.write('I1', 'WORKING DAYS IN MONTH')
    sheet_name.write('J4', 'CC Inventory ' + dateto)

    if sheet_name.get_name() == "CP":
        start2 = datetime.datetime(pomdate.year, pomdate.month, 1)
        end2 = datetime.datetime(pomdate.year, pomdate.month, calendar.monthrange(pomdate.year, pomdate.month)[1])
        spom2 = kon.working_days(start2, end2)
        sheet_name.write('L1', spom2)
    else:
        sheet_name.write_formula('L1', '=CP!L1')

    sheet_name.write('I2', 'WORKING DAYS PASSED')

    if sheet_name.get_name() == "CP":
        start3 = datetime.datetime(pomdate.year, pomdate.month, 1)
        end3 = datetime.datetime(pomdate.year, pomdate.month, int(kon.left(dateto, 2)))
        spom3 = kon.daysPassed(start3, end3)

        sheet_name.write('L2', spom3)
    else:
        sheet_name.write_formula('L2', '=CP!L2')

    sheet_name.write('A4', 'AWT code', cell_wrap)
    sheet_name.write('C4', 'Status SKU', cell_wrap)
    sheet_name.write('D4', 'Units / carton', cell_wrap)
    sheet_name.write('E4', 'Cartons / pallet', cell_wrap)
    sheet_name.write('I4', 'LAST 3 MONTHS AVG SALES', cell_wrap)

    # cell_green.set_font_color('green')
    # cell_green.set_text_wrap()
    # cell_green.set_align('center')
    # cell_green.set_align('vcenter')
    sheet_name.write('J4', 'CC Inventory ' + dateto, cell_green)

    # cell_blue.set_font_color('blue')
    # cell_blue.set_text_wrap()
    # cell_blue.set_align('center')
    # cell_blue.set_align('vcenter')
    sheet_name.write('K4', 'CS Inventory ' + dateto, cell_blue)

    # cell_red.set_font_color('red')
    # cell_red.set_text_wrap()
    # cell_red.set_align('center')
    # cell_red.set_align('vcenter')
    sheet_name.write('L4', 'Paid Inventory ' + dateto, cell_red)

    sheet_name.write('M4', 'WS Inventory ' + dateto, cell_wrap)
    sheet_name.write('N4', 'SALES TO TRADE UNITS', cell_wrap)
    sheet_name.write('O4', '% total sales', cell_wrap)
    sheet_name.write('P4', 'TIME GONE', cell_wrap)
    sheet_name.write('Q4', 'Linear projection for current month', cell_wrap)
    sheet_name.write('R4', 'Linear projection / Last 3 Month Average', cell_wrap)
    sheet_name.write('S4', 'Ratio STT / Last 3 Month Average', cell_wrap)
    sheet_name.write('T4', 'Stock in pallets', cell_wrap)
    sheet_name.write('U4', '% art / all stock', cell_wrap)
    sheet_name.write('V4', 'WDOC', cell_wrap)
    sheet_name.write('W4', 'WDOC WS', cell_wrap)
    sheet_name.write('X4', 'Expected WDOC', cell_wrap)
    sheet_name.write('Y4', 'Nabavna cena', cell_wrap)
    sheet_name.write('Z4', 'Nabavna Vrednost', cell_wrap)

    cell_percent.set_num_format('0.00%')
    sheet_name.write_formula('Q2', '=L2/L1', cell_percent)

    row = 4
    col = 0
    for i in range(26):
        sheet_name.write(row, col, i + 1, cell_wrap)
        col += 1

    sheet_name.set_column('A:A', 8.57)
    sheet_name.set_column('B:B', 60.43)
    sheet_name.set_column('C:C', 10.86)
    sheet_name.set_column('D:E', 8.14)
    sheet_name.set_column('F:H', 8.43)
    sheet_name.set_column('I:M', 9.14)
    sheet_name.set_column('N:P', 8.43)
    sheet_name.set_column('Q:R', 9.14)
    sheet_name.set_column('S:S', 9.10)
    sheet_name.set_column('T:U', 8.43)
    sheet_name.set_column('V:W', 8)
    sheet_name.set_column('X:X', 9.50)
    sheet_name.set_column('Y:Y', 7.61)
    sheet_name.set_column('Z:Z', 10.30)


def populateProductsSTT(sheet_name, vr_sifra, vr_naziv):
    totcnt = 0
    totals = [0]
    sumstart = [0]
    currcode = ""
    # aktivan = ""
    # GenNaz = ""
    lista_art = {}

    Imared = False

    nCnt = 5
    pomtot = ""
    sumstart[0] = 6
    # kartona_na_paleti = 0.0

    # cell_num.set_num_format('#,##0')
    # cell_tnum.set_num_format('#,##0')
    # cell_tnum.set_bg_color('yellow')
    #
    # cell_percent.set_num_format('0.00%')
    # cell_tpercent.set_num_format('0.00%')
    # cell_tpercent.set_bg_color('yellow')
    #
    # cell_tyellow.set_bg_color('yellow')

    sqltext = "SELECT ap.sifra as sifra,ap.naziv as naziv,ap.aktivan as aktivan, "
    sqltext = sqltext + "ap.produkt_gen_naziv_id as produkt_gen_naziv_id,"
    sqltext = sqltext + "(SELECT adm.pretvorba FROM adm_pakiranja adm, adm_produkti adm_prod, "
    sqltext = sqltext + "adm_produkti_u_pakiranjima adm_produp "
    sqltext = sqltext + "WHERE adm.ID = adm_produp.pakiranje_id AND (adm_prod.ID = adm_produp.produkt_id) "
    sqltext = sqltext + "AND (adm_prod.sifra = ap.sifra) AND adm.naziv like 'PALETA%' ) as paleta , "
    sqltext = sqltext + "(SELECT adm.pretvorba FROM adm_pakiranja adm,adm_produkti adm_prod, "
    sqltext = sqltext + "adm_produkti_u_pakiranjima adm_produp WHERE adm.ID = adm_produp.pakiranje_id "
    sqltext = sqltext + "AND (adm_prod.ID = adm_produp.produkt_id) AND (adm_prod.sifra = ap.sifra) "
    sqltext = sqltext + "AND adm.naziv like 'KARTON%' ) as karton "
    sqltext = sqltext + "FROM ADM_PRODUKTI ap WHERE SIFRA LIKE '" + vr_sifra + "%' AND aktivan = 'D' "
    sqltext = sqltext + "ORDER BY ap.SIFRA ASC"

    connectString = 'Driver={Oracle in OraClient10g_home1};DBQ=awt;UID=awtread;PWD=awtread'

    myconnection = pyodbc.connect(connectString)
    cursor = myconnection.cursor()
    cursor.execute(sqltext)
    data = cursor.fetchall()

    for polje in data:
        sifra = polje[0]
        naziv = polje[1]
        aktivan = polje[2]
        gen_naz_id = polje[3]
        paleta = polje[4]
        karton = polje[5]

        if kon.left(sifra, 4) != currcode:
            if (currcode != "") and Imared:
                if pomtot != "Total:":
                    # upisi total u redu
                    sheet_name.write(nCnt, 0, "Total:", cell_tyellow)

                    sheet_name.write(nCnt, 1, " ", cell_tyellow)
                    sheet_name.write(nCnt, 2, " ", cell_tyellow)
                    sheet_name.write(nCnt, 3, " ", cell_tyellow)
                    sheet_name.write(nCnt, 4, " ", cell_tyellow)

                    sheet_name.write_formula(nCnt, 5, "=SUM(F" + str(sumstart[totcnt]) + ":F" + str(nCnt) + ")",
                                             cell_tnum)
                    sheet_name.write_formula(nCnt, 6, "=SUM(G" + str(sumstart[totcnt]) + ":G" + str(nCnt) + ")",
                                             cell_tnum)
                    sheet_name.write_formula(nCnt, 7, "=SUM(H" + str(sumstart[totcnt]) + ":H" + str(nCnt) + ")",
                                             cell_tnum)

                    sheet_name.write_formula(nCnt, 8, "=AVERAGE(F" + str(nCnt + 1) + ":H" + str(nCnt + 1) + ")",
                                             cell_tnum)
                    sheet_name.write_formula(nCnt, 9, "=SUM(J" + str(sumstart[totcnt]) + ":J" + str(nCnt) + ")",
                                             cell_tnum)
                    sheet_name.write_formula(nCnt, 10, "=SUM(K" + str(sumstart[totcnt]) + ":K" + str(nCnt) + ")",
                                             cell_tnum)

                    sheet_name.write_formula(nCnt, 11, "=SUM(L" + str(sumstart[totcnt]) + ":L" + str(nCnt) + ")",
                                             cell_tnum)
                    sheet_name.write_formula(nCnt, 12, "=SUM(J" + str(nCnt + 1) + ":L" + str(nCnt + 1) + ")", cell_tnum)
                    sheet_name.write_formula(nCnt, 13, "=SUM(N" + str(sumstart[totcnt]) + ":N" + str(nCnt) + ")",
                                             cell_tnum)

                    sheet_name.write_formula(nCnt, 15, "=$Q$2", cell_tpercent)
                    sheet_name.write_formula(nCnt, 16,
                                             "=IF(ISERROR(N" + str(nCnt + 1) + "/P" + str(nCnt + 1) + "),0,(N" + str(
                                                 nCnt + 1) + "/P" + str(nCnt + 1) + "))", cell_tnum)

                    sheet_name.write_formula(nCnt, 17,
                                             "=IF(ISERROR(Q" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(Q" + str(
                                                 nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_tpercent)
                    sheet_name.write_formula(nCnt, 18,
                                             "=IF(ISERROR(N" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(N" + str(
                                                 nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_tpercent)

                    # kolona 20 ide posle grand totala
                    sheet_name.write_formula(nCnt, 19,
                                             "=IF(ISERROR(J" + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(
                                                 nCnt + 1) + ')," ",J' + str(nCnt + 1) + "/D" + str(
                                                 nCnt + 1) + "/E" + str(nCnt + 1) + ")", cell_tnum)
                    sheet_name.write_formula(nCnt, 21,
                                             "=IF(ISERROR(J" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,J" + str(
                                                 nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_tnum)
                    sheet_name.write_formula(nCnt, 22,
                                             "=IF(ISERROR(M" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,M" + str(
                                                 nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_tnum)
                    sheet_name.write_formula(nCnt, 23,
                                             "=IF(ISERROR((M" + str(nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
                                                 nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(
                                                 nCnt + 1) + ",G" + str(nCnt + 1) + ")*22),0,(M" + str(
                                                 nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
                                                 nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(
                                                 nCnt + 1) + ",G" + str(nCnt + 1) + ")*22)", cell_tnum)
                    # prazno polje za cenu i vrednost u totalu
                    sheet_name.write(nCnt, 24, 0, cell_tnum)
                    # sheet_name.write(nCnt, 25, 0,cell_tnum)
                    sheet_name.write_formula(nCnt, 25, "=SUM(Z" + str(sumstart[totcnt]) + ":Z" + str(nCnt) + ")",
                                             cell_tnum)

                    totcnt = totcnt + 1
                    nCnt = nCnt + 1

                    totals.append(nCnt)
                    sumstart.append(nCnt + 1)

                    Imared = False
                    pomtot = "Total:"

        if aktivan == "D":

            sheet_name.write(nCnt, 0, sifra)

            if gen_naz_id == 5545:
                GenNaz = "REDOVAN"
            elif gen_naz_id == 5546:
                GenNaz = "DELISTIRAN"
            elif gen_naz_id == 5547:
                GenNaz = "DELISTIRAN - ZAMENJEN"
            elif gen_naz_id == 5548:
                GenNaz = "PROMO"
            elif gen_naz_id == 5549:
                GenNaz = "PROMO - DELISTIRAN"
            elif gen_naz_id == 5550:
                GenNaz = "PROMO - ZAMENJEN"
            else:
                GenNaz = "NEMA"

            sheet_name.write(nCnt, 1, naziv)
            sheet_name.write(nCnt, 2, GenNaz)

            lista_art[sifra] = nCnt

            if karton:
                sheet_name.write(nCnt, 3, int(karton), cell_num)
                if paleta:
                    kartona_na_paleti = paleta / karton
                    sheet_name.write(nCnt, 4, int(kartona_na_paleti), cell_num)

            sheet_name.write(nCnt, 5, 0)
            sheet_name.write(nCnt, 6, 0)
            sheet_name.write(nCnt, 7, 0)

            sheet_name.write_formula(nCnt, 8, "=AVERAGE(F" + str(nCnt + 1) + ":H" + str(nCnt + 1) + ")", cell_num)
            sheet_name.write_formula(nCnt, 12, "=SUM(J" + str(nCnt + 1) + ":L" + str(nCnt + 1) + ")", cell_num)

            # kolona 14 ide kroz grand total
            sheet_name.write_formula(nCnt, 15, "=$Q$2", cell_percent)
            sheet_name.write_formula(nCnt, 16, "=IF(ISERROR(N" + str(nCnt + 1) + "/P" + str(nCnt + 1) + "),0,(N" + str(
                nCnt + 1) + "/P" + str(nCnt + 1) + "))", cell_num)

            sheet_name.write_formula(nCnt, 17, "=IF(ISERROR(Q" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(Q" + str(
                nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_percent)
            sheet_name.write_formula(nCnt, 18, "=IF(ISERROR(N" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(N" + str(
                nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_percent)

            sheet_name.write_formula(nCnt, 19, "=IF(ISERROR(J" + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(
                nCnt + 1) + ')," ",J' + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(nCnt + 1) + ")", cell_num)
            # NOVI WDOC
            sheet_name.write_formula(nCnt, 21,
                                     "=IF(ISERROR(J" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,J" + str(
                                         nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_num)
            # WDOC WS
            sheet_name.write_formula(nCnt, 22,
                                     "=IF(ISERROR(M" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,M" + str(
                                         nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_num)
            # Expected WDOC
            sheet_name.write_formula(nCnt, 23, "=IF(ISERROR((M" + str(nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
                nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(nCnt + 1) + ",G" + str(
                nCnt + 1) + ")*22),0,(M" + str(nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
                nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(nCnt + 1) + ",G" + str(nCnt + 1) + ")*22)",
                                     cell_num)
            sheet_name.write_formula(nCnt, 25,
                                     "=IF(ISERROR(M" + str(nCnt + 1) + "*Y" + str(nCnt + 1) + "),0,M" + str(
                                         nCnt + 1) + "*Y" + str(nCnt + 1) + ")", cell_num)
            nCnt = nCnt + 1
            Imared = True
            pomtot = ""

        currcode = kon.left(sifra, 4)

    if pomtot != "Total:":
        sheet_name.write(nCnt, 0, "Total:", cell_tyellow)

        sheet_name.write(nCnt, 1, " ", cell_tyellow)
        sheet_name.write(nCnt, 2, " ", cell_tyellow)
        sheet_name.write(nCnt, 3, " ", cell_tyellow)
        sheet_name.write(nCnt, 4, " ", cell_tyellow)

        sheet_name.write_formula(nCnt, 5, "=SUM(F" + str(sumstart[totcnt]) + ":F" + str(nCnt) + ")", cell_tnum)
        sheet_name.write_formula(nCnt, 6, "=SUM(G" + str(sumstart[totcnt]) + ":G" + str(nCnt) + ")", cell_tnum)
        sheet_name.write_formula(nCnt, 7, "=SUM(H" + str(sumstart[totcnt]) + ":H" + str(nCnt) + ")", cell_tnum)

        sheet_name.write_formula(nCnt, 8, "=AVERAGE(F" + str(nCnt + 1) + ":H" + str(nCnt + 1) + ")", cell_tnum)
        sheet_name.write_formula(nCnt, 9, "=SUM(J" + str(sumstart[totcnt]) + ":J" + str(nCnt) + ")", cell_tnum)
        sheet_name.write_formula(nCnt, 10, "=SUM(K" + str(sumstart[totcnt]) + ":K" + str(nCnt) + ")", cell_tnum)

        sheet_name.write_formula(nCnt, 11, "=SUM(L" + str(sumstart[totcnt]) + ":L" + str(nCnt) + ")", cell_tnum)
        sheet_name.write_formula(nCnt, 12, "=SUM(J" + str(nCnt + 1) + ":L" + str(nCnt + 1) + ")", cell_tnum)
        sheet_name.write_formula(nCnt, 13, "=SUM(N" + str(sumstart[totcnt]) + ":N" + str(nCnt) + ")", cell_tnum)

        # kolona 14 ide kroz grandtotal
        sheet_name.write_formula(nCnt, 15, "=$Q$2", cell_tpercent)
        sheet_name.write_formula(nCnt, 16, "=IF(ISERROR(N" + str(nCnt + 1) + "/P" + str(nCnt + 1) + "),0,(N" + str(
            nCnt + 1) + "/P" + str(nCnt + 1) + "))", cell_tnum)

        sheet_name.write_formula(nCnt, 17, "=IF(ISERROR(Q" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(Q" + str(
            nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_tpercent)
        sheet_name.write_formula(nCnt, 18, "=IF(ISERROR(N" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(N" + str(
            nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_tpercent)
        sheet_name.write_formula(nCnt, 19, "=IF(ISERROR(J" + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(
            nCnt + 1) + ')," ",J' + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(nCnt + 1) + ")", cell_tnum)

        sheet_name.write_formula(nCnt, 21, "=IF(ISERROR(J" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,J" + str(
            nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_tnum)
        sheet_name.write_formula(nCnt, 22, "=IF(ISERROR(M" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,M" + str(
            nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_tnum)
        sheet_name.write_formula(nCnt, 23, "=IF(ISERROR((M" + str(nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
            nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(nCnt + 1) + ",G" + str(
            nCnt + 1) + ")*22),0,(M" + str(nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
            nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(nCnt + 1) + ",G" + str(nCnt + 1) + ")*22)",
                                 cell_tnum)

        # prazno polje za cenu i vrednost u totalu
        sheet_name.write(nCnt, 24, 0, cell_tnum)
        # sheet_name.write(nCnt, 25, 0, cell_tnum)
        sheet_name.write_formula(nCnt, 25, "=SUM(Z" + str(sumstart[totcnt]) + ":Z" + str(nCnt) + ")", cell_tnum)

        totcnt = totcnt + 1
        nCnt = nCnt + 1
        totals.append(nCnt)
        sumstart.append(nCnt + 1)
    else:
        nCnt = nCnt - 1

    sheet_name.write(nCnt, 0, "Total " + vr_naziv)

    for j in range(5, nCnt + 1):
        sheet_name.write_formula(j, 14, "=IF(ISERROR(I" + str(j + 1) + "/$I$" + str(totals[-1] + 1) + "),0,I" + str(
            j + 1) + "/$I$" + str(totals[-1] + 1) + ")", cell_percent)
        sheet_name.write_formula(j, 20, "=IF(ISERROR(T" + str(j + 1) + "/$T$" + str(totals[-1] + 1) + "),0,T" + str(
            j + 1) + "/$T$" + str(totals[-1] + 1) + ")", cell_percent)

    totalrange = "="
    for i in range(1, totcnt + 1):
        totalrange = totalrange + "F" + str(totals[i]) + "+"
        sheet_name.write_formula(int(str(totals[i])) - 1, 14,
                                 "=IF(ISERROR(I" + str(totals[i]) + "/$I$" + str(totals[-1] + 1) + "),0,I" + str(
                                     totals[i]) + "/$I$" + str(totals[-1] + 1) + ")", cell_tpercent)
        sheet_name.write_formula(int(str(totals[i])) - 1, 20,
                                 "=IF(ISERROR(T" + str(totals[i]) + "/$T$" + str(totals[-1] + 1) + "),0,T" + str(
                                     totals[i]) + "/$T$" + str(totals[-1] + 1) + ")", cell_tpercent)

    totalrange = kon.left(totalrange, len(totalrange) - 1)
    sheet_name.write_formula(nCnt, 5, totalrange)

    totalrange = totalrange.replace("F", "G")
    sheet_name.write_formula(nCnt, 6, totalrange, cell_num)

    totalrange = totalrange.replace("G", "H")
    sheet_name.write_formula(nCnt, 7, totalrange, cell_num)

    sheet_name.write_formula(nCnt, 8, "=AVERAGE(F" + str(nCnt + 1) + ":H" + str(nCnt + 1) + ")", cell_num)

    totalrange = totalrange.replace("H", "J")
    sheet_name.write_formula(nCnt, 9, totalrange, cell_num)

    totalrange = totalrange.replace("J", "K")
    sheet_name.write_formula(nCnt, 10, totalrange, cell_num)

    totalrange = totalrange.replace("K", "L")
    sheet_name.write_formula(nCnt, 11, totalrange, cell_num)

    sheet_name.write_formula(nCnt, 12, "=SUM(J" + str(nCnt + 1) + ":L" + str(nCnt + 1) + ")", cell_num)

    totalrange = totalrange.replace("L", "N")
    sheet_name.write_formula(nCnt, 13, totalrange, cell_num)

    sheet_name.write_formula(nCnt, 14, "=IF(ISERROR(I" + str(nCnt + 1) + "/$I$" + str(nCnt + 1) + "),0,I" + str(
        nCnt + 1) + "/$I$" + str(nCnt + 1) + ")", cell_percent)
    sheet_name.write_formula(nCnt, 15, "=$Q$2", cell_percent)
    sheet_name.write_formula(nCnt, 16, "=IF(ISERROR(N" + str(nCnt + 1) + "/P" + str(nCnt + 1) + "),0,(N" + str(
        nCnt + 1) + "/P" + str(nCnt + 1) + "))", cell_num)

    sheet_name.write_formula(nCnt, 17, "=IF(ISERROR(Q" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(Q" + str(
        nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_percent)
    sheet_name.write_formula(nCnt, 18, "=IF(ISERROR(N" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "),0,(N" + str(
        nCnt + 1) + "/I" + str(nCnt + 1) + "))", cell_percent)
    sheet_name.write_formula(nCnt, 19, "=IF(ISERROR(J" + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(
        nCnt + 1) + ')," ",J' + str(nCnt + 1) + "/D" + str(nCnt + 1) + "/E" + str(nCnt + 1) + ")", cell_num)

    sheet_name.write_formula(nCnt, 21, "=IF(ISERROR(J" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,J" + str(
        nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_num)
    sheet_name.write_formula(nCnt, 22, "=IF(ISERROR(M" + str(nCnt + 1) + "/I" + str(nCnt + 1) + "*22),0,M" + str(
        nCnt + 1) + "/I" + str(nCnt + 1) + "*22)", cell_num)
    sheet_name.write_formula(nCnt, 23, "=IF(ISERROR((M" + str(nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(
        nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(nCnt + 1) + ",G" + str(nCnt + 1) + ")*22),0,(M" + str(
        nCnt + 1) + "-(Q" + str(nCnt + 1) + "-N" + str(nCnt + 1) + "))/AVERAGE(Q" + str(nCnt + 1) + ",H" + str(
        nCnt + 1) + ",G" + str(nCnt + 1) + ")*22)", cell_num)

    totalrange = totalrange.replace("N", "Z")
    sheet_name.write_formula(nCnt, 25, totalrange, cell_num)

    cursor.close()
    myconnection.close()

    del cursor
    del myconnection

    return lista_art


def readSLV(slv_file, cs_kol, cc_kol, paid_kol):
    book = opx.load_workbook(slv_file, data_only=True)
    sheets = book.sheetnames
    lista_slv = []
    for sheet_name in sheets:
        if sheet_name == "SLV Total":
            sh = book.active
            maxrow = sh.max_row
            for row in sh.iter_rows(min_row=2, max_row=maxrow, min_col=1, max_col=int(paid_kol), values_only=True):
                cell1 = row[1]
                cell2 = row[int(cc_kol) - 1]
                cell3 = row[int(cs_kol) - 1]
                cell4 = row[int(paid_kol) - 1]
                # print(cell1, cell2, cell3, cell4) test sta vraca
                lista_slv.append((cell1, cell2, cell3, cell4))
    book.close()
    del book
    return lista_slv


def checkSLV(slv_file):
    book = opx.load_workbook(slv_file)
    sheets = book.sheetnames
    ok_slv = 1
    for sheet_name in sheets:
        if sheet_name == "SLV Total":
            ok_slv = 0
    book.close()
    del book
    return ok_slv


def populateSTTFromSLV(sheet_name, ar_lista, slv_lista):
    # red = 0
    for key in ar_lista.keys():
        for polje in slv_lista:
            sifra = polje[0]
            cc_col = polje[1]
            cs_col = polje[2]
            paid_col = polje[3]
            if sifra == key:
                red = ar_lista.get(key, 0)
                sheet_name.write(red, 9, cc_col)
                sheet_name.write(red, 10, cs_col)
                sheet_name.write(red, 11, paid_col)
                break


def populateThreeMonths(sheet_name, vr_sifra, ar_lista, dateto):
    # red = 0
    # suma = 0
    # pomsum = 0

    pomdate = datetime.datetime.now()

    # mesDo0 = calendar.monthrange(pomdate.year, pomdate.month)[1]

    if pomdate.month == 1:
        mesDo1 = calendar.monthrange(pomdate.year - 1, 12)[1]
        mesDo2 = calendar.monthrange(pomdate.year - 1, 11)[1]
        mesDo3 = calendar.monthrange(pomdate.year - 1, 10)[1]

    elif pomdate.month == 2:
        mesDo1 = calendar.monthrange(pomdate.year, pomdate.month - 1)[1]
        mesDo2 = calendar.monthrange(pomdate.year - 1, 12)[1]
        mesDo3 = calendar.monthrange(pomdate.year - 1, 11)[1]

    elif pomdate.month == 3:
        mesDo1 = calendar.monthrange(pomdate.year, pomdate.month - 1)[1]
        mesDo2 = calendar.monthrange(pomdate.year, pomdate.month - 2)[1]
        mesDo3 = calendar.monthrange(pomdate.year - 1, 12)[1]

    else:
        mesDo1 = calendar.monthrange(pomdate.year, pomdate.month - 1)[1]
        mesDo2 = calendar.monthrange(pomdate.year, pomdate.month - 2)[1]
        mesDo3 = calendar.monthrange(pomdate.year, pomdate.month - 3)[1]

    # tekuci mesec sales to trade vrednost

    date_od = datetime.datetime(pomdate.year, pomdate.month, 1)
    date_od1 = date_od.strftime("%d.%m.%Y")

    # date_do = dateto
    # date_do = datetime.datetime(pomdate.year,pomdate.month, mesDo0 )
    # date_do1 = date_do.strftime("%d.%m.%Y")
    date_do1 = dateto

    sqltext = "SELECT ADM_PRODUKTI.SIFRA as sifra3, sum(ROBSTAV.IZLAZNA_KOLICINA) as sum3 "
    sqltext = sqltext + "FROM ADM_PARTNERI ADM, ADM_SUBJEKTI ADMSUBJ, ROB_SKL_DOKUMENTI_ZAG ROB,"
    sqltext = sqltext + " ROB_SKL_DOKUMENTI_STAV ROBSTAV,ADM_PRODUKTI "
    sqltext = sqltext + "WHERE ( SUBSTR(ADM_PRODUKTI.SIFRA,1,2)= '" + vr_sifra + "' AND "
    sqltext = sqltext + "(ADM.ID = ADMSUBJ.PARTNER_ID) AND (ADMSUBJ.ID = ROB.SUBJEKT_ID) AND "
    sqltext = sqltext + "(ADMSUBJ.ID = ROBSTAV.R_SUBJEKT_ID) AND (ROB.ID = ROBSTAV.SKL_DOKUMENT_ZAG_ID) AND "
    sqltext = sqltext + "(ROB.DATUM_DVO >= '" + date_od1 + "') AND "
    sqltext = sqltext + "(ROB.DATUM_DVO <= '" + date_do1 + "') AND "
    sqltext = sqltext + "(ADM_PRODUKTI.ID = ROBSTAV.PRODUKT_ID) AND (ROB.VRSTA_TRANSAKCIJE_ID <> 9) AND "
    sqltext = sqltext + "(ADM_PRODUKTI.AKTIVAN = 'D') and (ROBSTAV.IZLAZNA_KOLICINA<>0)) "
    sqltext = sqltext + "group by ROBSTAV.PRODUKT_ID,ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV "
    sqltext = sqltext + "order by ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV"

    connectString = 'Driver={Oracle in OraClient10g_home1};DBQ=awt;UID=awtread;PWD=awtread'
    myconnection = pyodbc.connect(connectString)
    cursor = myconnection.cursor()
    cursor.execute(sqltext)
    data = cursor.fetchall()

    for polje in data:
        sifra = polje[0]
        pomsum = polje[1]
        suma = round(pomsum)

        for key in ar_lista.keys():
            if sifra == key:
                red = ar_lista.get(key, 0)
                sheet_name.write(red, 13, suma)
                break

    cursor.close()
    myconnection.close()
    #    del cursor
    #    del myconnection

    # prvi mesec
    if pomdate.month == 1:
        date_od = datetime.datetime(pomdate.year - 1, 12, 1)
    else:
        date_od = datetime.datetime(pomdate.year, pomdate.month - 1, 1)
    date_od1 = date_od.strftime("%d.%m.%Y")

    if pomdate.month == 1:
        date_do = datetime.datetime(pomdate.year - 1, 12, 31)
    else:
        date_do = datetime.datetime(pomdate.year, pomdate.month - 1, mesDo1)
    date_do1 = date_do.strftime("%d.%m.%Y")

    sqltext = "SELECT ADM_PRODUKTI.SIFRA as sifra3, sum(ROBSTAV.IZLAZNA_KOLICINA) as sum3 "
    sqltext = sqltext + "FROM ADM_PARTNERI ADM, ADM_SUBJEKTI ADMSUBJ, ROB_SKL_DOKUMENTI_ZAG ROB, "
    sqltext = sqltext + "ROB_SKL_DOKUMENTI_STAV ROBSTAV,ADM_PRODUKTI "
    sqltext = sqltext + "WHERE ( SUBSTR(ADM_PRODUKTI.SIFRA,1,2)= '" + vr_sifra + "' AND "
    sqltext = sqltext + "(ADM.ID = ADMSUBJ.PARTNER_ID) AND (ADMSUBJ.ID = ROB.SUBJEKT_ID) AND "
    sqltext = sqltext + "(ADMSUBJ.ID = ROBSTAV.R_SUBJEKT_ID) AND (ROB.ID = ROBSTAV.SKL_DOKUMENT_ZAG_ID) AND "
    sqltext = sqltext + "(ROB.DATUM_DVO >= '" + date_od1 + "') AND "
    sqltext = sqltext + "(ROB.DATUM_DVO <= '" + date_do1 + "') AND "
    sqltext = sqltext + "(ADM_PRODUKTI.ID = ROBSTAV.PRODUKT_ID) AND (ROB.VRSTA_TRANSAKCIJE_ID <> 9) AND "
    sqltext = sqltext + "(ADM_PRODUKTI.AKTIVAN = 'D') and (ROBSTAV.IZLAZNA_KOLICINA<>0)) "
    sqltext = sqltext + "group by ROBSTAV.PRODUKT_ID,ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV "
    sqltext = sqltext + "order by ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV"

    connectString = 'Driver={Oracle in OraClient10g_home1};DBQ=awt;UID=awtread;PWD=awtread'
    myconnection = pyodbc.connect(connectString)
    cursor = myconnection.cursor()
    cursor.execute(sqltext)
    data = cursor.fetchall()

    for polje in data:
        sifra = polje[0]
        pomsum = polje[1]
        suma = round(pomsum)

        for key in ar_lista.keys():
            if sifra == key:
                red = ar_lista.get(key, 0)
                sheet_name.write(red, 7, suma)
                break

    cursor.close()
    myconnection.close()
    #    del cursor
    #    del myconnection

    # drugi mesec

    if pomdate.month == 1:
        date_od = datetime.datetime(pomdate.year - 1, 11, 1)
    elif pomdate.month == 2:
        date_od = datetime.datetime(pomdate.year - 1, 12, 1)
    else:
        date_od = datetime.datetime(pomdate.year, pomdate.month - 2, 1)
    date_od1 = date_od.strftime("%d.%m.%Y")

    if pomdate.month == 1:
        date_do = datetime.datetime(pomdate.year - 1, 11, 30)
    elif pomdate.month == 2:
        date_do = datetime.datetime(pomdate.year - 1, 12, 31)
    else:
        date_do = datetime.datetime(pomdate.year, pomdate.month - 2, mesDo2)
    date_do1 = date_do.strftime("%d.%m.%Y")

    sqltext = "SELECT ADM_PRODUKTI.SIFRA as sifra3, sum(ROBSTAV.IZLAZNA_KOLICINA) as sum3 "
    sqltext = sqltext + "FROM ADM_PARTNERI ADM, ADM_SUBJEKTI ADMSUBJ, ROB_SKL_DOKUMENTI_ZAG ROB, "
    sqltext = sqltext + "ROB_SKL_DOKUMENTI_STAV ROBSTAV,ADM_PRODUKTI "
    sqltext = sqltext + "WHERE ( SUBSTR(ADM_PRODUKTI.SIFRA,1,2)= '" + vr_sifra + "' AND "
    sqltext = sqltext + "(ADM.ID = ADMSUBJ.PARTNER_ID) AND (ADMSUBJ.ID = ROB.SUBJEKT_ID) AND "
    sqltext = sqltext + "(ADMSUBJ.ID = ROBSTAV.R_SUBJEKT_ID) AND (ROB.ID = ROBSTAV.SKL_DOKUMENT_ZAG_ID) AND "
    sqltext = sqltext + "(ROB.DATUM_DVO >= '" + date_od1 + "') AND "
    sqltext = sqltext + "(ROB.DATUM_DVO <= '" + date_do1 + "') AND "
    sqltext = sqltext + "(ADM_PRODUKTI.ID = ROBSTAV.PRODUKT_ID) AND (ROB.VRSTA_TRANSAKCIJE_ID <> 9) AND "
    sqltext = sqltext + "(ADM_PRODUKTI.AKTIVAN = 'D') and (ROBSTAV.IZLAZNA_KOLICINA<>0)) "
    sqltext = sqltext + "group by ROBSTAV.PRODUKT_ID,ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV "
    sqltext = sqltext + "order by ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV"

    connectString = 'Driver={Oracle in OraClient10g_home1};DBQ=awt;UID=awtread;PWD=awtread'
    myconnection = pyodbc.connect(connectString)
    cursor = myconnection.cursor()
    cursor.execute(sqltext)
    data = cursor.fetchall()

    for polje in data:
        sifra = polje[0]
        pomsum = polje[1]
        suma = round(pomsum)

        for key in ar_lista.keys():
            if sifra == key:
                red = ar_lista.get(key, 0)
                sheet_name.write(red, 6, suma)
                break
    cursor.close()
    myconnection.close()

    # treci mesec

    if pomdate.month == 1:
        date_od = datetime.datetime(pomdate.year - 1, 10, 1)
    elif pomdate.month == 2:
        date_od = datetime.datetime(pomdate.year - 1, 11, 1)
    elif pomdate.month == 3:
        date_od = datetime.datetime(pomdate.year - 1, 12, 1)
    else:
        date_od = datetime.datetime(pomdate.year, pomdate.month - 3, 1)
    date_od1 = date_od.strftime("%d.%m.%Y")

    if pomdate.month == 1:
        date_do = datetime.datetime(pomdate.year - 1, 10, 31)
    elif pomdate.month == 2:
        date_do = datetime.datetime(pomdate.year - 1, 11, 30)
    elif pomdate.month == 3:
        date_do = datetime.datetime(pomdate.year - 1, 12, 31)
    else:
        date_do = datetime.datetime(pomdate.year, pomdate.month - 3, mesDo3)
    date_do1 = date_do.strftime("%d.%m.%Y")

    sqltext = "SELECT ADM_PRODUKTI.SIFRA as sifra3, sum(ROBSTAV.IZLAZNA_KOLICINA) as sum3 "
    sqltext = sqltext + "FROM ADM_PARTNERI ADM, ADM_SUBJEKTI ADMSUBJ, ROB_SKL_DOKUMENTI_ZAG ROB, "
    sqltext = sqltext + "ROB_SKL_DOKUMENTI_STAV ROBSTAV,ADM_PRODUKTI "
    sqltext = sqltext + "WHERE ( SUBSTR(ADM_PRODUKTI.SIFRA,1,2)= '" + vr_sifra + "' AND "
    sqltext = sqltext + "(ADM.ID = ADMSUBJ.PARTNER_ID) AND (ADMSUBJ.ID = ROB.SUBJEKT_ID) AND "
    sqltext = sqltext + "(ADMSUBJ.ID = ROBSTAV.R_SUBJEKT_ID) AND (ROB.ID = ROBSTAV.SKL_DOKUMENT_ZAG_ID) AND "
    sqltext = sqltext + "(ROB.DATUM_DVO >= '" + date_od1 + "') AND "
    sqltext = sqltext + "(ROB.DATUM_DVO <= '" + date_do1 + "') AND "
    sqltext = sqltext + "(ADM_PRODUKTI.ID = ROBSTAV.PRODUKT_ID) AND (ROB.VRSTA_TRANSAKCIJE_ID <> 9) AND "
    sqltext = sqltext + "(ADM_PRODUKTI.AKTIVAN = 'D') and (ROBSTAV.IZLAZNA_KOLICINA<>0)) "
    sqltext = sqltext + "group by ROBSTAV.PRODUKT_ID,ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV "
    sqltext = sqltext + "order by ADM_PRODUKTI.SIFRA, ADM_PRODUKTI.NAZIV"

    connectString = 'Driver={Oracle in OraClient10g_home1};DBQ=awt;UID=awtread;PWD=awtread'
    myconnection = pyodbc.connect(connectString)
    cursor = myconnection.cursor()
    cursor.execute(sqltext)
    data = cursor.fetchall()

    for polje in data:
        sifra = polje[0]
        pomsum = polje[1]
        suma = round(pomsum)

        for key in ar_lista.keys():
            if sifra == key:
                red = ar_lista.get(key, 0)
                sheet_name.write(red, 5, suma)
                break
    cursor.close()
    myconnection.close()
    del cursor
    del myconnection


def populateNabavna(sheet_name, vr_sifra, ar_lista):
    # red = 0
    # ncena = 0
    # pomsum = 0
    # nabavna cena

    sqltext = "SELECT P.SIFRA AS sifra, ROUND(SUM(Round(Z.KNJIZENA_KOLICINA * C.NABAVNA_CIJENA, 2)) /"
    sqltext = sqltext + " Sum(Z.KNJIZENA_KOLICINA), 2) AS nc "
    sqltext = sqltext + "FROM ROB_SKL_ZALIHE Z, ROB_SKL_CIJENE C, ADM_PRODUKTI P "
    sqltext = sqltext + "WHERE (Z.PRODUKT_ID = C.PRODUKT_ID) AND (Z.SKLADISTE_ID = C.SKLADISTE_ID) "
    sqltext = sqltext + "AND (Z.PRODUKT_ID = P.ID) "
    sqltext = sqltext + "AND (Z.KNJIZENA_KOLICINA > 0) AND ( SUBSTR(P.SIFRA,1,2)= '" + vr_sifra + "')"
    sqltext = sqltext + "GROUP BY P.SIFRA "
    sqltext = sqltext + "ORDER BY P.SIFRA ASC"

    connectString = 'Driver={Oracle in OraClient10g_home1};DBQ=awt;UID=awtread;PWD=awtread'
    myconnection = pyodbc.connect(connectString)
    cursor = myconnection.cursor()
    cursor.execute(sqltext)
    data = cursor.fetchall()

    for polje in data:
        sifra = polje[0]
        ncena = polje[1]

        for key in ar_lista.keys():
            if sifra == key:
                red = ar_lista.get(key, 0)
                sheet_name.write(red, 24, ncena)
                break

    cursor.close()
    myconnection.close()
    del cursor
    del myconnection


def generateExcel(dateto_text, slv_file, cs_kol, cc_kol, paid_kol, textinfo, bar):
    # Create an new Excel file and add a worksheet.
    sttname = 'Z:/STT Report {:%Y-%m-%d %H %M}.xlsx'.format(datetime.datetime.now())

    global cell_red
    global cell_blue
    global cell_green
    global cell_yellow
    global cell_percent
    global cell_wrap
    global cell_num
    global cell_tnum
    global cell_tpercent
    global cell_tyellow

    workbook = xlsxwriter.Workbook(sttname)

    cell_red = workbook.add_format()
    cell_red.set_font_color('red')
    cell_red.set_text_wrap()
    cell_red.set_align('center')
    cell_red.set_align('vcenter')

    cell_blue = workbook.add_format()
    cell_blue.set_font_color('blue')
    cell_blue.set_text_wrap()
    cell_blue.set_align('center')
    cell_blue.set_align('vcenter')

    cell_green = workbook.add_format()
    cell_green.set_font_color('green')
    cell_green.set_text_wrap()
    cell_green.set_align('center')
    cell_green.set_align('vcenter')

    cell_yellow = workbook.add_format()

    cell_percent = workbook.add_format()
    cell_percent.set_num_format('0.00%')

    cell_num = workbook.add_format()
    cell_num.set_num_format('#,##0')

    # cell_tpercent = workbook.add_format()
    # cell_tpercent.set_num_format('0.00%')
    # cell_tpercent.set_bg_color('yellow')

    cell_tpercent = workbook.add_format({'num_format': '0.00%', 'bg_color': 'yellow'})
    # cell_tpercent.set_num_format('0.00%')
    # cell_tpercent.set_bg_color('yellow')

    cell_tnum = workbook.add_format()
    cell_tnum.set_num_format('#,##0')
    cell_tnum.set_bg_color('yellow')

    cell_tyellow = workbook.add_format({'font_color': 'yellow'})
    # cell_tyellow.set_bg_color('yellow')

    cell_wrap = workbook.add_format()
    cell_wrap.set_text_wrap()
    cell_wrap.set_align('center')
    cell_wrap.set_align('vcenter')

    lista_art = {}
    # slv_art = []

    vrste = (["CP", "01", 1], ["Filiz", "10", 2], ["Philips", "11", 3], ["Barilla", "12", 4], ["Naturel", "18", 5],
             ["Naturel Musli", "19", 6],
             ["Heinz", "23", 7], ["Weaver Popcorn", "21", 8], ["Naturel Krekeri", "30", 9], ["Zewa", "33", 10],
             ["Libresse", "37", 11],
             ["Kotanyi", "06", 12], ["Naturel Pastete", "35", 13], ["Labud", "39", 14], ["Aroma", "41", 15],
             ["Erdal", "42", 16], ["Franck", "51", 17], ["Philadelphia", "55", 18], ["Juicy", "54", 19],
             ["FS Savex", "57", 20], ["FS Semana", "58", 21], ["FS Pufies", "59", 22], ["Ostalo", "99", 23])

    # Ocitaj slv i ubaci u niz
    slv_art = readSLV(slv_file, cs_kol, cc_kol, paid_kol)

    for vrsta, sifra, red in vrste:
        if red != 24:
            worksheet = workbook.add_worksheet(vrsta)
            worksheet.set_landscape()

            kon.updateVrsta(vrsta, textinfo)

            kon.updateBar(red, bar)

            populateSTTSheet(worksheet, dateto_text)

            lista_art = populateProductsSTT(worksheet, sifra, vrsta)

            populateSTTFromSLV(worksheet, lista_art, slv_art)

            populateThreeMonths(worksheet, sifra, lista_art, dateto_text)

            populateNabavna(worksheet, sifra, lista_art)

    workbook.close()


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print('PyCharm')
